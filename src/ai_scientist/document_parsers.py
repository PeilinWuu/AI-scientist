"""Bounded local parsers for user-provided research documents and datasets."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import zipfile
from collections import Counter
from pathlib import Path
from typing import Any

from src.ai_scientist.schemas import ParsedAssetContent


MAX_EXTRACTED_CHARS = int(os.getenv("AI_SCIENTIST_ASSET_MAX_EXTRACTED_CHARS", "50000"))
MAX_TABLE_ROWS = int(os.getenv("AI_SCIENTIST_ASSET_MAX_TABLE_ROWS", "500"))
MAX_SAMPLE_ROWS = int(os.getenv("AI_SCIENTIST_ASSET_SAMPLE_ROWS", "10"))
MAX_PDF_PAGES = int(os.getenv("AI_SCIENTIST_ASSET_MAX_PDF_PAGES", "100"))
MAX_XML_NODES = int(os.getenv("AI_SCIENTIST_ASSET_MAX_XML_NODES", "5000"))
MAX_XLSX_UNCOMPRESSED_BYTES = int(
    os.getenv("AI_SCIENTIST_ASSET_MAX_XLSX_UNCOMPRESSED_BYTES", str(100 * 1024 * 1024))
)


def parse_research_asset(path: Path) -> ParsedAssetContent:
    """Parse one supported file into a bounded, model-safe structured representation."""

    suffix = path.suffix.lower()
    digest = _sha256(path)
    if suffix in {".txt", ".md"}:
        return _parse_text(path, digest)
    if suffix in {".csv", ".tsv"}:
        return _parse_delimited(path, digest, suffix)
    if suffix == ".json":
        return _parse_json(path, digest)
    if suffix == ".xml":
        return _parse_xml(path, digest)
    if suffix == ".pdf":
        return _parse_pdf(path, digest)
    if suffix == ".xlsx":
        return _parse_xlsx(path, digest)
    if suffix == ".xls":
        return _parse_xls(path, digest)
    raise ValueError(f"No local parser is registered for {suffix or 'this file type'}.")


def _parse_text(path: Path, digest: str) -> ParsedAssetContent:
    text, encoding = _read_text(path)
    excerpt, truncated = _bounded(text)
    lines = text.splitlines()
    return ParsedAssetContent(
        parser_name="local_text_v1",
        content_kind="document",
        summary=f"Text document with {len(lines)} lines and {len(text)} characters.",
        extracted_text=excerpt,
        structured_summary={"encoding": encoding, "line_count": len(lines), "character_count": len(text)},
        content_sha256=digest,
        truncated=truncated,
    )


def _parse_delimited(path: Path, digest: str, suffix: str) -> ParsedAssetContent:
    text, encoding = _read_text(path)
    sample = text[:8192]
    fallback_delimiter = "\t" if suffix == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = fallback_delimiter
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[list[str]] = []
    scan_limit_reached = False
    for index, row in enumerate(reader):
        if index > MAX_TABLE_ROWS:
            scan_limit_reached = True
            break
        rows.append([str(value) for value in row])
    if not rows:
        raise ValueError("Delimited data file contains no rows.")
    headers = _unique_headers(rows[0])
    data_rows = rows[1:]
    missing = {
        header: sum(1 for row in data_rows if index >= len(row) or not row[index].strip())
        for index, header in enumerate(headers)
    }
    samples = [_row_record(headers, row) for row in data_rows[:MAX_SAMPLE_ROWS]]
    partial = scan_limit_reached
    summary = {
        "encoding": encoding,
        "delimiter": "tab" if delimiter == "\t" else delimiter,
        "column_names": headers,
        "scanned_data_rows": len(data_rows),
        "row_count_is_partial": partial,
        "missing_values_in_scanned_rows": missing,
        "sample_rows": samples,
    }
    extracted, truncated = _bounded(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return ParsedAssetContent(
        parser_name="local_delimited_v1",
        content_kind="tabular",
        summary=f"Tabular data with {len(headers)} columns and {len(data_rows)} scanned data rows.",
        extracted_text=extracted,
        structured_summary=summary,
        warnings=["Row count is partial because the parser scan limit was reached."] if partial else [],
        content_sha256=digest,
        truncated=truncated,
    )


def _parse_json(path: Path, digest: str) -> ParsedAssetContent:
    text, encoding = _read_text(path)
    value = json.loads(text)
    summary: dict[str, Any] = {"encoding": encoding, "root_type": type(value).__name__}
    if isinstance(value, dict):
        summary["top_level_keys"] = list(value)[:100]
        summary["top_level_key_count"] = len(value)
    elif isinstance(value, list):
        summary["item_count"] = len(value)
        summary["sample_items"] = value[:MAX_SAMPLE_ROWS]
        if value and all(isinstance(item, dict) for item in value[:MAX_SAMPLE_ROWS]):
            summary["observed_fields"] = sorted({key for item in value[:MAX_SAMPLE_ROWS] for key in item})
    pretty = json.dumps(value, ensure_ascii=False, indent=2, default=str)
    excerpt, truncated = _bounded(pretty)
    return ParsedAssetContent(
        parser_name="local_json_v1",
        content_kind="structured_data",
        summary=f"JSON document with root type {summary['root_type']}.",
        extracted_text=excerpt,
        structured_summary=summary,
        content_sha256=digest,
        truncated=truncated,
    )


def _parse_xml(path: Path, digest: str) -> ParsedAssetContent:
    try:
        from defusedxml import ElementTree
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("XML parsing requires defusedxml.") from exc
    root = ElementTree.parse(path).getroot()
    tag_counts: Counter[str] = Counter()
    text_fragments: list[str] = []
    scanned = 0
    for node in root.iter():
        scanned += 1
        if scanned > MAX_XML_NODES:
            break
        tag = _local_xml_name(str(node.tag))
        tag_counts[tag] += 1
        value = " ".join((node.text or "").split())
        if value:
            text_fragments.append(f"{tag}: {value}")
    excerpt, truncated_by_chars = _bounded("\n".join(text_fragments))
    node_truncated = scanned > MAX_XML_NODES
    summary = {
        "root_tag": _local_xml_name(str(root.tag)),
        "scanned_node_count": min(scanned, MAX_XML_NODES),
        "top_tags": tag_counts.most_common(50),
    }
    return ParsedAssetContent(
        parser_name="local_defusedxml_v1",
        content_kind="structured_data",
        summary=f"XML document rooted at {summary['root_tag']} with {summary['scanned_node_count']} scanned nodes.",
        extracted_text=excerpt,
        structured_summary=summary,
        warnings=["XML node scan limit reached."] if node_truncated else [],
        content_sha256=digest,
        truncated=truncated_by_chars or node_truncated,
    )


def _parse_pdf(path: Path, digest: str) -> ParsedAssetContent:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("PDF parsing requires pypdf.") from exc
    reader = PdfReader(str(path))
    if reader.is_encrypted:
        try:
            reader.decrypt("")
        except Exception as exc:
            raise ValueError("Encrypted PDF cannot be parsed without a password.") from exc
    page_count = len(reader.pages)
    page_limit = min(page_count, MAX_PDF_PAGES)
    fragments: list[str] = []
    empty_pages = 0
    for index in range(page_limit):
        text = (reader.pages[index].extract_text() or "").strip()
        if not text:
            empty_pages += 1
            continue
        fragments.append(f"[Page {index + 1}]\n{text}")
    excerpt, char_truncated = _bounded("\n\n".join(fragments))
    warnings = []
    if page_count > page_limit:
        warnings.append("PDF page limit reached.")
    if empty_pages:
        warnings.append(f"{empty_pages} scanned pages contained no extractable text; OCR may be required.")
    return ParsedAssetContent(
        parser_name="local_pypdf_v1",
        content_kind="document",
        summary=f"PDF with {page_count} pages; text extracted from {page_limit - empty_pages} scanned pages.",
        extracted_text=excerpt,
        structured_summary={
            "page_count": page_count,
            "scanned_page_count": page_limit,
            "empty_text_page_count": empty_pages,
            "metadata": {str(key): str(value) for key, value in (reader.metadata or {}).items()},
        },
        warnings=warnings,
        content_sha256=digest,
        truncated=char_truncated or page_count > page_limit,
    )


def _parse_xlsx(path: Path, digest: str) -> ParsedAssetContent:
    _validate_xlsx_archive(path)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("XLSX parsing requires openpyxl.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    sheets: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        for worksheet in workbook.worksheets[:20]:
            rows: list[list[Any]] = []
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index > MAX_TABLE_ROWS:
                    warnings.append(f"Sheet {worksheet.title}: row scan limit reached.")
                    break
                rows.append([_cell_value(value) for value in row])
            if rows:
                headers = _unique_headers(rows[0])
                samples = [_row_record(headers, row) for row in rows[1 : MAX_SAMPLE_ROWS + 1]]
            else:
                headers, samples = [], []
            sheets.append(
                {
                    "sheet_name": worksheet.title,
                    "column_names": headers,
                    "scanned_data_rows": max(0, len(rows) - 1),
                    "sample_rows": samples,
                }
            )
    finally:
        workbook.close()
    summary = {"sheet_names": sheet_names, "sheets": sheets}
    extracted, truncated = _bounded(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return ParsedAssetContent(
        parser_name="local_openpyxl_v1",
        content_kind="tabular",
        summary=f"Excel workbook with {len(sheet_names)} worksheets.",
        extracted_text=extracted,
        structured_summary=summary,
        warnings=warnings,
        content_sha256=digest,
        truncated=truncated or bool(warnings),
    )


def _parse_xls(path: Path, digest: str) -> ParsedAssetContent:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("XLS parsing requires xlrd.") from exc
    workbook = xlrd.open_workbook(path, on_demand=True)
    sheet_names = list(workbook.sheet_names())
    sheets: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        for name in sheet_names[:20]:
            sheet = workbook.sheet_by_name(name)
            scan_rows = min(sheet.nrows, MAX_TABLE_ROWS + 1)
            rows = [[_cell_value(value) for value in sheet.row_values(index)] for index in range(scan_rows)]
            headers = _unique_headers(rows[0]) if rows else []
            sheets.append(
                {
                    "sheet_name": name,
                    "column_names": headers,
                    "total_data_rows": max(0, sheet.nrows - 1),
                    "sample_rows": [_row_record(headers, row) for row in rows[1 : MAX_SAMPLE_ROWS + 1]],
                }
            )
            if sheet.nrows > scan_rows:
                warnings.append(f"Sheet {name}: row scan limit reached.")
    finally:
        workbook.release_resources()
    summary = {"sheet_names": sheet_names, "sheets": sheets}
    extracted, truncated = _bounded(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return ParsedAssetContent(
        parser_name="local_xlrd_v1",
        content_kind="tabular",
        summary=f"Legacy Excel workbook with {len(summary['sheet_names'])} worksheets.",
        extracted_text=extracted,
        structured_summary=summary,
        warnings=warnings,
        content_sha256=digest,
        truncated=truncated or bool(warnings),
    )


def _validate_xlsx_archive(path: Path) -> None:
    with zipfile.ZipFile(path) as archive:
        total = sum(item.file_size for item in archive.infolist())
        if total > MAX_XLSX_UNCOMPRESSED_BYTES:
            raise ValueError("XLSX expanded content exceeds the configured safety limit.")


def _read_text(path: Path) -> tuple[str, str]:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "utf-8-replacement"


def _bounded(text: str) -> tuple[str, bool]:
    normalized = text.replace("\x00", "")
    if len(normalized) <= MAX_EXTRACTED_CHARS:
        return normalized, False
    return normalized[:MAX_EXTRACTED_CHARS], True


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _unique_headers(row: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: Counter[str] = Counter()
    for index, value in enumerate(row):
        base = str(value).strip() or f"column_{index + 1}"
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def _row_record(headers: list[str], row: list[Any]) -> dict[str, Any]:
    return {header: _cell_value(row[index]) if index < len(row) else None for index, header in enumerate(headers)}


def _cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _local_xml_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]
